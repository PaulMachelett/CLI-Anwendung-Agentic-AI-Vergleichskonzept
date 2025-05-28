import sys
from collections import defaultdict
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from Prompts_Data import prompts, always_asked_criteria




# Agentenname eingeben
agentenname = input("Bitte gib den Namen des Agenten ein: ").strip()

agenten_slug = agentenname.strip().lower().replace(" ", "_")

# Ergebnisse speichern
results = {}
kriterien_textantworten = {}
extra_prompts_gesamt = 0


# Immer abgefragte Kriterien nach allen Prompts separat abfragen
print("\n### Abfrage immer geltender Kriterien ###\n")
results["Always Asked"] = []

# Bewertungsprozess
for prompt, criteria in prompts.items():
    print(f"\n{prompt} - Bitte bewerte die folgenden Kriterien:\n")
    results[prompt] = []

    # Frage nach Anzahl extra Prompts – außer bei Prompt 13
    if prompt != "Prompt 13":
        while True:
            try:
                anzahl_extra = int(input("Wie viele extra Prompts wurden benötigt bis der Code vollständig war und funktionierte? "))
                if anzahl_extra >= 0:
                    extra_prompts_gesamt += anzahl_extra
                    break
                else:
                    print("Bitte eine nicht-negative Zahl eingeben.")
            except ValueError:
                print("Ungültige Eingabe. Bitte eine Zahl eingeben.")


    # Normale Kriterien
    for idx, criterion in enumerate(criteria):
        print("\n" + "#" * 30 + "\n")
        print(f"Kriterium {idx + 1}: {criterion['name']}")
        print(f"Beschreibung: {criterion['description']}")
        for i, option in enumerate(criterion['options'], 1):
            print(f"  [{i}] {option}")
        print()

        while True:
            try:
                choice_input = input("Bitte Nummer der Bewertung eingeben: ")
                choice = int(choice_input)
                if 1 <= choice <= len(criterion['options']):

                    bewertungstext = criterion['options'][choice - 1]

                    if len(criterion['options']) == 2:
                        score = 1.0 if choice == 1 else 0.0
                    elif len(criterion['options']) == 3:
                        score = [1.0, 0.5, 0.0][choice - 1]
                    elif len(criterion['options']) == 4:
                        score = [1.0, 0.66, 0.33, 0.0][choice - 1]
                    else:
                        score = 0.0
                    break
                else:
                    print("Ungültige Eingabe.")
            except ValueError:
                print("Bitte eine gültige Zahl eingeben.")

        results[prompt].append({
            "name": criterion['name'],
            "category": criterion['category'],
            "score": score
        })

        if criterion['name'] not in kriterien_textantworten:
            kriterien_textantworten[criterion['name']] = bewertungstext


    # 👇 Hier: Nur wenn NICHT Prompt 13 → Immer-gefragt-Kriterien abfragen
    if prompt != "Prompt 13":
        for idx, criterion in enumerate(always_asked_criteria):
            print("\n" + "#" * 30 + "\n")
            print(f"(Immer gefragt) Kriterium {idx + 1}: {criterion['name']}")
            print(f"Beschreibung: {criterion['description']}")
            for i, option in enumerate(criterion['options'], 1):
                print(f"  [{i}] {option}")
            print()

            while True:
                try:
                    choice_input = input("Bitte Nummer der Bewertung eingeben (oder Enter für keine Bewertung): ")
                    if choice_input.strip() == "":
                        score = None
                        break
                    choice = int(choice_input)
                    if 1 <= choice <= len(criterion['options']):
                        if len(criterion['options']) == 2:
                            score = 1.0 if choice == 1 else 0.0
                        elif len(criterion['options']) == 3:
                            score = [1.0, 0.5, 0.0][choice - 1]
                        elif len(criterion['options']) == 4:
                            score = [1.0, 0.66, 0.33, 0.0][choice - 1]
                        else:
                            score = 0.0
                        break
                    else:
                        print("Ungültige Eingabe.")
                except ValueError:
                    print("Bitte eine gültige Zahl eingeben.")

            if score is not None:
                results[prompt].append({
                    "name": criterion['name'],
                    "category": criterion['category'],
                    "score": score
                })

    # Eingabe zur Fortsetzung
    while True:
        cont = input("Möchtest du mit dem nächsten Prompt fortfahren? (j/n): ").strip().lower()
        if cont == 'j':
            break
        elif cont == 'n':
            prompt_abbruch = True
            break
        else:
            print("Falsche Eingabe. Bitte j für Ja oder n für Nein.")
    if cont == 'n':
        break
    
# Auswertung    
kriterium_mittelwerte = defaultdict(list)
kriterium_kategorie_map = {}

for prompt_criteria in results.values():
    for criterion in prompt_criteria:
        key = criterion['name'].strip()
        kriterium_mittelwerte[key].append(criterion['score'])
        kriterium_kategorie_map[key] = criterion['category']

# --- Neue Code-Kriterien einmalig abfragen ---
code_kriterien_einmalig = [
    {
        "name": "Code Security",
        "description": "Überprüft die Anzahl der gefundenen Sicherheitslücken.",
        "category": "Code Kriterium"
    },
    {
        "name": "Code Reliability",
        "description": "Überprüft, wie viele Bugs innerhalb des Codes vorhanden sind.",
        "category": "Code Kriterium"
    },
    {
        "name": "Code Maintainability",
        "description": "Untersucht, wie gut sich der erzeugte Code verstehen, erweitern und pflegen lässt.",
        "category": "Code Kriterium"
    }
]

print(f"\n\n#### Gesamtanzahl benötigter extra Prompts: {extra_prompts_gesamt} ####")

# Werte aus einmaliger Abfrage speichern
code_kriterien_wert = {}

print("\nSonarQube Wert der folgenden Code Kriterien:")

for criterion in code_kriterien_einmalig:
    while True:
        try:
            print(f"\n{criterion['name']}: {criterion['description']}")
            value = float(input("Bitte Zahl eingeben: "))
            if value >= 0:
                code_kriterien_wert[criterion["name"]] = {
                    "category": criterion["category"],
                    "score": value
                }
                break
            else:
                print("Bitte eine positive Zahl eingeben.")
        except ValueError:
            print("Ungültige Eingabe. Bitte eine gültige Zahl eingeben.")

# --- Mittelwerte für bestehende "Code Kriterium"-Kriterien berechnen ---
code_kriterien_mittelwerte = {}

for criterion_name, scores in kriterium_mittelwerte.items():
    if kriterium_kategorie_map[criterion_name] == "Code Kriterium":
        mean_score = sum(scores) / len(scores)
        code_kriterien_mittelwerte[criterion_name] = {
            "category": "Code Kriterium",
            "score": mean_score
        }

# --- Ergänze die einmaligen numerischen Werte ---
code_kriterien_mittelwerte.update(code_kriterien_wert)



# --- Ausgabe aller "Code Kriterium"-Werte ---
print("\n\n\n#### Bewertungen Kriterien der Kategorie 'Code Kriterium': ####")
for name, info in code_kriterien_mittelwerte.items():
    if name in ["Code Security", "Code Reliability", "Code Maintainability"]:
        print(f"- {name}: {int(info['score'])} SonarQube Meldung")
    else:
        print(f"- {name}: {int(info['score'] * 100)} % der Fälle")

# Kategorie-Mittelwerte berechnen (ohne "Code Kriterium")
final_category_averages = defaultdict(list)
for criterion_name, scores in kriterium_mittelwerte.items():
    category = kriterium_kategorie_map[criterion_name]
    if category == "Code Kriterium":
        continue
    mean_crit = sum(scores) / len(scores)
    final_category_averages[category].append(mean_crit)

# Alle Kategorien ermitteln (auch unbeurteilte anzeigen)
all_categories = set(criterion['category'] for criteria in prompts.values() for criterion in criteria)
all_categories.update(c['category'] for c in always_asked_criteria)

print("\n\n#### Durchschnittsbewertungen Kategorien: ####")
for category in sorted(all_categories):
    if category == "Code Kriterium":
        continue
    values = final_category_averages.get(category, [])
    if values:
        mean = sum(values) / len(values)
        print(f"- {category}: {mean:.2f}")
    else:
        print(f"- {category}: n/a")

# Autonomie-Mittelwert berechnen
autonomie_quellen = ["Lernfähigkeit", "Memory", "Planning", "Reflexion", "Systemintegration", "Tool-Usage"]
autonomie_werte = [
    sum(final_category_averages[cat]) / len(final_category_averages[cat])
    for cat in autonomie_quellen if cat in final_category_averages and final_category_averages[cat]
]
autonomie_mittelwert = sum(autonomie_werte) / len(autonomie_werte) if autonomie_werte else 0.0

# Radar-Diagramm Werte korrekt berechnen
agenten_kategorien = ["Pro-Aktivität", "Kommunikation", "Reaktivität", "Coding Practices", "Systemintegration", "Autonomie"]
agenten_werte = []
for cat in agenten_kategorien:
    if cat == "Autonomie":
        agenten_werte.append(autonomie_mittelwert)
    else:
        werte = final_category_averages.get(cat, [])
        wert = sum(werte) / len(werte) if werte else 0
        agenten_werte.append(wert)

# Autonomie-Radar
autonomie_kategorien = ["Lernfähigkeit", "Memory", "Planning", "Reflexion", "Tool-Usage"]
autonomie_detail_werte = []
for cat in autonomie_kategorien:
    werte = final_category_averages.get(cat, [])
    wert = sum(werte) / len(werte) if werte else 0
    autonomie_detail_werte.append(wert)

def zeichne_radar(titel, kategorien, werte, dateiname):
    winkel = np.linspace(0, 2 * np.pi, len(kategorien), endpoint=False).tolist()
    werte += werte[:1]
    winkel += winkel[:1]

    fig, ax = plt.subplots(figsize=(6, 6), subplot_kw=dict(polar=True))
    ax.plot(winkel, werte, marker='o')
    ax.fill(winkel, werte, alpha=0.25)
    ax.set_title(titel, size=14, y=1.1)
    ax.set_xticks(winkel[:-1])
    ax.set_xticklabels(kategorien)
    ax.set_yticks([0.2, 0.4, 0.6, 0.8])
    ax.set_yticklabels(["0.2", "0.4", "0.6", "0.8"])
    ax.set_ylim(0, 1)

    plt.tight_layout()
    plt.savefig(dateiname)
    plt.close()

# Radar-Grafiken speichern
# Datei-Namen dynamisch auf Basis des Agentennamens
datei_eigenschaften = f"{agenten_slug}_eigenschaften.png"
datei_autonomie = f"{agenten_slug}_autonomie.png"

# Radar-Grafiken speichern
zeichne_radar(f"{agentenname} Eigenschaften", agenten_kategorien, agenten_werte.copy(), datei_eigenschaften)
zeichne_radar(f"{agentenname} Autonomie", autonomie_kategorien, autonomie_detail_werte.copy(), datei_autonomie)

print(f"\nRadar-Diagramme wurden gespeichert als '{datei_eigenschaften}' und '{datei_autonomie}' für Agent: {agentenname}")

# --- Excel-Datei mit Bewertungstexten erstellen ---
excel_data = []

for kriterium, bewertung in kriterien_textantworten.items():
    excel_data.append({
        "Kriterium": kriterium,
        "Bewertung": bewertung
    })

# Daten vorbereiten
data = list(kriterien_textantworten.items())
df = pd.DataFrame(data, columns=["Kriterium", "Bewertung"])

# Datei schreiben
excel_filename = f"{agenten_slug}_kriterien_textantworten.xlsx"
df.to_excel(excel_filename, index=False)

# Spaltenbreite automatisch an längsten Text anpassen
wb = load_workbook(excel_filename)
ws = wb.active

# Berechne maximale Textlänge aus allen Zellen beider Spalten
all_lengths = [
    len(str(cell.value)) for col in ws.columns for cell in col if cell.value
]
max_length = max(all_lengths) if all_lengths else 20
adjusted_width = max_length + 2  # etwas Puffer

# Setze gleiche Breite für beide Spalten
for col_letter in ["A", "B"]:
    ws.column_dimensions[col_letter].width = adjusted_width


wb.save(excel_filename)

print(f"\nExcel-Datei mit Bewertungstexten gespeichert als '{excel_filename}'")

