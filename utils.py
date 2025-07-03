import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
import pandas as pd
from collections import defaultdict
from Data import (
    always_asked_criteria,
    prompts,
    sekundäre_faktoren,
    autonomie_quellen,
    agenten_kategorien,
    prompt_mapping,
)
from Prompts import prompts_text


def bewertungs_prozess(
    kriterien_textantworten,
    extra_prompts_gesamt,
    agentenname,
    sekundäre_kriterien_werte,
    agenten_werte,
    autonomie_detail_werte,
    code_kriterien_wert,
):
    results = {}  # Speichert für jeden Prompt Kriterium, Kategorie sowie Bewertungs des Kriteriums
    code_kriterien_werte = {"Ausführbarer Code": 0, "Anforderungserfüllung": 0}

    print(f"\nBewertung des Agenten: {agentenname}\n")

    for prompt_key, criteria in prompts.items():
        prompt_text_key = prompt_mapping.get(prompt_key)
        prompt_text = prompts_text.get(prompt_text_key, "")

        print(f"\nBeschreibung für: {prompt_key}\n\n{prompt_text.strip()}\n")
        print(f"\nBitte bewerte die folgenden Kriterien:\n")

        results[prompt_key] = []

        # Normale Kriterien
        for idx, criterion in enumerate(criteria):
            # print("\n" + "#" * 30 + "\n")
            print(f"Kriterium {idx + 1}: {criterion['name']}")
            print(f"Beschreibung: {criterion['description']}")
            for i, option in enumerate(criterion["options"], 1):
                print(f"  [{i}] {option}")
            print()

            while True:
                try:
                    choice_input = input("Bitte Nummer der Bewertung eingeben: ")
                    choice = int(choice_input)
                    if 1 <= choice <= len(criterion["options"]):
                        bewertungstext = criterion["options"][choice - 1]

                        if len(criterion["options"]) == 2:
                            score = 1.0 if choice == 1 else 0.0
                        elif len(criterion["options"]) == 3:
                            score = [1.0, 0.5, 0.0][choice - 1]
                        elif len(criterion["options"]) == 4:
                            score = [1.0, 0.66, 0.33, 0.0][choice - 1]
                        else:
                            score = 0.0
                        break
                    else:
                        print("Ungültige Eingabe.")
                except ValueError:
                    print("Bitte eine gültige Zahl eingeben.")

            results[prompt_key].append(
                {
                    "name": criterion["name"],
                    "category": criterion["category"],
                    "score": score,
                }
            )

            if criterion["name"] not in kriterien_textantworten:
                kriterien_textantworten[criterion["name"]] = bewertungstext

        # Anzahl extra Prompts
        if prompt_key != "Agenten Features":
            while True:
                try:
                    anzahl_extra_prompts = int(
                        input(
                            "Wie viele extra Prompts wurden benötigt bis der Code vollständig war und funktionierte? "
                        )
                    )
                    if anzahl_extra_prompts >= 0:
                        extra_prompts_gesamt[0] += anzahl_extra_prompts
                        break

                    else:
                        print("Bitte eine nicht-negative Zahl eingeben.")
                except ValueError:
                    print("Ungültige Eingabe. Bitte eine Zahl eingeben.")

        # Immer-gefragt-Kriterien
        if prompt_key != "Agenten Features":
            for idx, criterion in enumerate(always_asked_criteria):
                print("\n" + "#" * 30 + "\n")
                print(f"(Immer gefragt) Kriterium {idx + 1}: {criterion['name']}")
                print(f"Beschreibung: {criterion['description']}")
                for i, option in enumerate(criterion["options"], 1):
                    print(f"  [{i}] {option}")
                print()

                while True:
                    try:
                        if criterion["name"] in [
                            "Ausführbarer Code",
                            "Anforderungserfüllung",
                        ]:
                            choice_input = input(
                                "Bitte Nummer der Bewertung eingeben: "
                            ).strip()
                            if choice_input == "":
                                print(
                                    "Für dieses Kriterium ist eine Bewertung erforderlich."
                                )
                                continue
                        else:
                            choice_input = input(
                                "Bitte Nummer der Bewertung eingeben (oder Enter für keine Bewertung): "
                            ).strip()
                            if choice_input == "":
                                score = None
                                break
                        choice = int(choice_input)
                        if 1 <= choice <= len(criterion["options"]):
                            if len(criterion["options"]) == 2:
                                score = 1.0 if choice == 1 else 0.0
                            elif len(criterion["options"]) == 3:
                                score = [1.0, 0.5, 0.0][choice - 1]
                            elif len(criterion["options"]) == 4:
                                score = [1.0, 0.66, 0.33, 0.0][choice - 1]
                            else:
                                score = 0.0
                            break
                        else:
                            print("Ungültige Eingabe.")
                    except ValueError:
                        print("Bitte eine gültige Zahl eingeben.")

                if score is not None:
                    if criterion["name"] in [
                        "Ausführbarer Code",
                        "Anforderungserfüllung",
                    ]:
                        code_kriterien_werte[criterion["name"]] += score

                    else:
                        results[prompt_key].append(
                            {
                                "name": criterion["name"],
                                "category": criterion["category"],
                                "score": score,
                            }
                        )

        # Eingabe zur Fortsetzung
        while True:
            cont = (
                input("Möchtest du mit dem nächsten Prompt fortfahren? (j/n): ")
                .strip()
                .lower()
            )
            if cont == "j":
                break
            elif cont == "n":
                break
            else:
                print("Falsche Eingabe. Bitte j für Ja oder n für Nein.")
        if cont == "n":
            break

    process_bewertungen(
        results,
        sekundäre_kriterien_werte,
        agenten_werte,
        autonomie_detail_werte,
        code_kriterien_werte,
        code_kriterien_wert,
    )


def process_bewertungen(
    results,
    sekundäre_kriterien_werte,
    agenten_werte,
    autonomie_detail_werte,
    code_kriterien_werte,
    code_kriterien_wert,
):
    # Speichert als Key Kriterienname und als Value eine Liste mit allen Bewertungen dieses Kriteriums
    kriterium_bewertungen = defaultdict(list)

    # Speichert alle Mittelwerte der Kriterien einer Kategorie, WICHTIG nicht die Mittelwerte der Kategorie, das kommt erst später ---
    # Sprich dieses Dict enthält als Key die Kategorie und dann dann als Value eine Liste mit allen Mittelwerten der Kriterien dieser Kategorie ---
    kriterien_mittelwerte_per_katergorie = defaultdict(list)

    # Speichert für jedes Kriterium(key) die Kategorie(value)
    kriterium_name_kategorie_map = {}

    # Einträge aus Results in "kriterium_bewertungen" überführt
    for prompt_criteria in results.values():
        for criterion in prompt_criteria:
            key = criterion["name"].strip()
            kriterium_bewertungen[key].append(criterion["score"])
            kriterium_name_kategorie_map[key] = criterion["category"]

    # Fragt sekunäre Kriterien ab und speichert diese in "sekundäre_kriterien_werte"
    input_sek_kriterien(sekundäre_kriterien_werte, sekundäre_faktoren)

    # Kategorie-Mittelwerte berechnen
    for criterion_name, scores in kriterium_bewertungen.items():
        category = kriterium_name_kategorie_map[criterion_name]
        if category == "Code Kriterium":
            continue
        mean_crit = sum(scores) / len(scores)
        kriterien_mittelwerte_per_katergorie[category].append(mean_crit)

    # Mittelwert der Autonomie Kategorien berechnen
    autonomie_werte = [
        sum(kriterien_mittelwerte_per_katergorie[cat])
        / len(kriterien_mittelwerte_per_katergorie[cat])
        for cat in autonomie_quellen
        if cat in kriterien_mittelwerte_per_katergorie
        and kriterien_mittelwerte_per_katergorie[cat]
    ]

    # Finalen Autonomie Mittelwert berechnen
    autonomie_mittelwert = (
        sum(autonomie_werte) / len(autonomie_werte) if autonomie_werte else 0.0
    )

    # Finale Kategorie Mittelwerte berechnen
    for cat in agenten_kategorien:
        if cat == "Autonomie":
            agenten_werte.append(autonomie_mittelwert)
        else:
            werte = kriterien_mittelwerte_per_katergorie.get(cat, [])
            wert = sum(werte) / len(werte) if werte else 0
            agenten_werte.append(wert)

    # Autonomie
    for cat in autonomie_quellen:
        werte = kriterien_mittelwerte_per_katergorie.get(cat, [])
        wert = sum(werte) / len(werte) if werte else 0
        autonomie_detail_werte.append(wert)

    berechne_code_kriterien_mittelwert(code_kriterien_werte, code_kriterien_wert)


def input_sek_kriterien(sekundäre_kriterien_werte, sekundäre_faktoren):
    print("\n\nBewertung der sekundären Kriterien\n\n")

    for idx, criterion in enumerate(sekundäre_faktoren):
        print(f"Sekundäres Kriterium {idx + 1}: {criterion['name']}")
        print(f"Beschreibung: {criterion['description']}")
        for i, option in enumerate(criterion["options"], 1):
            print(f"  [{i}] {option}")
        print()

        while True:
            try:
                choice_input = input(
                    f"Bitte Nummer der Bewertung eingeben (1–{len(criterion['options'])}): "
                )
                if choice_input.strip() == "":
                    print(
                        "Leere Eingabe ist nicht erlaubt. Bitte eine gültige Zahl eingeben."
                    )
                    continue
                choice = int(choice_input)
                if 1 <= choice <= len(criterion["options"]):
                    if len(criterion["options"]) == 2:
                        score = 1.0 if choice == 1 else 0.0
                    elif len(criterion["options"]) == 3:
                        score = [1.0, 0.5, 0.0][choice - 1]
                    elif len(criterion["options"]) == 4:
                        score = [1.0, 0.66, 0.33, 0.0][choice - 1]
                    else:
                        score = 0.0  # Fallback bei ungewöhnlicher Optionenzahl
                    break
                else:
                    print(
                        f"Ungültige Eingabe. Bitte eine Zahl zwischen 1 und {len(criterion['options'])} eingeben."
                    )
            except ValueError:
                print("Ungültige Eingabe. Bitte eine gültige Zahl eingeben.")

        print("\n" + "#" * 30 + "\n")

        # Bewertung speichern
        sekundäre_kriterien_werte["kriterium_name"].append(criterion["name"])
        sekundäre_kriterien_werte["werte"].append(score)


def berechne_code_kriterien_mittelwert(code_kriterien_werte, code_kriterien_wert):
    for key, wert in code_kriterien_werte.items():
        if wert:  # Nur wenn Liste nicht leer ist
            prozent_zahl = wert / 11  # 11 ist die Prompt Anzahl
            code_kriterien_wert[key] = prozent_zahl
        else:
            code_kriterien_wert[key] = 0  # Optional: None für leere Listen


def zeichne_balkendiagramm(
    titel, kategorien, werte1, werte2, label1, label2, dateiname
):
    x = np.arange(len(kategorien))
    breite = 0.35

    fig, ax = plt.subplots(figsize=(12, 6))

    balken1 = ax.bar(x - breite / 2, werte1, breite, label=label1)
    balken2 = ax.bar(x + breite / 2, werte2, breite, label=label2)

    ax.set_title(titel, fontsize=14, y=1.05)
    ax.set_xlabel("Kategorien")
    ax.set_ylabel("Werte")
    ax.set_xticks(x)
    ax.set_xticklabels(
        kategorien, rotation=45, ha="right"
    )  # ⬅️ Rotation hilft gegen Überlappung
    ax.set_ylim(0, 1)
    ax.legend()

    for bars in [balken1, balken2]:
        for bar in bars:
            yval = bar.get_height()
            ax.text(
                bar.get_x() + bar.get_width() / 2,
                yval + 0.02,
                f"{yval:.2f}",
                ha="center",
                va="bottom",
                fontsize=8,
            )

    plt.tight_layout(rect=[0, 0.15, 1, 0.95])  # ⬅️ mehr Platz unten
    plt.savefig(dateiname)
    plt.close()


def create_excel_table(
    kriterien_textantworten1,
    agenten_slug1,
    extra_prompts_gesamt1,
    sekundäre_kriterien_werte1,
    code_kriterien_wert1,
    kriterien_textantworten2,
    agenten_slug2,
    extra_prompts_gesamt2,
    sekundäre_kriterien_werte2,
    code_kriterien_wert2,
):
    # --- Falls Slugs gleich sind, umbenennen zur Sicherheit ---
    if agenten_slug1 == agenten_slug2:
        agenten_slug1 += "_1"
        agenten_slug2 += "_2"

    # --- Textantworten in DataFrames ---
    df1 = pd.DataFrame(
        list(kriterien_textantworten1.items()), columns=["Kriterium", agenten_slug1]
    )
    df2 = pd.DataFrame(
        list(kriterien_textantworten2.items()), columns=["Kriterium", agenten_slug2]
    )
    df = pd.merge(df1, df2, on="Kriterium", how="outer")

    # --- Leerzeile vorbereiten ---
    df_empty = pd.DataFrame(
        columns=["Kriterium", agenten_slug1, agenten_slug2], data=[["", "", ""]]
    )

    # --- Alle Abschnitte zusammenführen ---
    df_final = pd.concat([df, df_empty], ignore_index=True)

    # --- Excel-Datei schreiben ---
    excel_filename = f"{agenten_slug1}_vs_{agenten_slug2}_kriterien.xlsx"
    df_final.to_excel(excel_filename, index=False)

    # --- Openpyxl für Feinschliff ---
    wb = load_workbook(excel_filename)
    ws = wb.active

    # Spaltenbreite automatisch setzen
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    # Leerzeile + extra Prompts am Ende
    last_row = ws.max_row + 2
    ws.cell(row=last_row, column=1, value="Anzahl benötigter extra Prompts")
    ws.cell(row=last_row, column=2, value=extra_prompts_gesamt1)
    ws.cell(row=last_row, column=3, value=extra_prompts_gesamt2)

    # Neue Zeile: Ausführbarer Code und Anforderungserfüllung
    last_row += 2
    ws.cell(row=last_row, column=1, value="Ausführbarer Code")
    ws.cell(
        row=last_row,
        column=2,
        value=f"{code_kriterien_wert1['Ausführbarer Code'] * 100:.0f} %",
    )
    ws.cell(
        row=last_row,
        column=3,
        value=f"{code_kriterien_wert2['Ausführbarer Code'] * 100:.0f} %",
    )
    last_row += 1
    ws.cell(row=last_row, column=1, value="Anforderungserfüllung")
    ws.cell(
        row=last_row,
        column=2,
        value=f"{code_kriterien_wert1['Anforderungserfüllung'] * 100:.0f} %",
    )
    ws.cell(
        row=last_row,
        column=3,
        value=f"{code_kriterien_wert2['Anforderungserfüllung'] * 100:.0f} %",
    )

    last_row = ws.max_row + 2
    ws.cell(row=last_row, column=1, value="Preisgestaltung")
    ws.cell(
        row=last_row,
        column=2,
        value=get_price_description(sekundäre_kriterien_werte1["werte"][-1]),
    )
    ws.cell(
        row=last_row,
        column=3,
        value=get_price_description(sekundäre_kriterien_werte2["werte"][-1]),
    )

    # Neue Zeile: Open/Closed Source
    last_row += 1
    ws.cell(row=last_row, column=1, value="Open/Closed Source")
    ws.cell(
        row=last_row,
        column=2,
        value=get_openclosedsource_description(sekundäre_kriterien_werte1["werte"][-2]),
    )
    ws.cell(
        row=last_row,
        column=3,
        value=get_openclosedsource_description(sekundäre_kriterien_werte2["werte"][-2]),
    )

    wb.save(excel_filename)
    print(f"\n✅ Excel-Datei gespeichert als '{excel_filename}'")


def get_price_description(rating):
    if rating == 1:
        return "Agent sowie Model (lokal) komplett kostenlos"
    elif rating == 0.5:
        return "Agent kostenlos, nur Zahlung nach API Key Nutzung"
    elif rating == 0:
        return "Festgelegter Preis pro Monat (inkl. Agent und API Key Nutzung)."
    else:
        return "Keine Angabe"


def get_openclosedsource_description(rating):
    if rating == 1:
        return "open source"
    elif rating == 0:
        return "closed source"
    else:
        return "Keine Angabe"
